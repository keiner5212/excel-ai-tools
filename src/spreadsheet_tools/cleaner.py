from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_tools.utils import index_to_column, normalize_scalar


def is_empty_value(value: object | None) -> bool:
    return normalize_scalar(value) is None


def row_has_content(values: list[object | None]) -> bool:
    return any(not is_empty_value(value) for value in values)


def trim_trailing_empty_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    trimmed = list(rows)
    while trimmed and not row_has_content(
        [cell.get("value") for cell in trimmed[-1]["cells"]]
    ):
        trimmed.pop()
    return trimmed


def _col_has_content_or_master(rows: list[dict[str, Any]], col_idx: int) -> bool:
    """Return True if the column has a non-empty value OR a merge master cell.

    Merge master cells are writable even when currently empty; trimming them
    would hide valid write targets from the AI.
    """
    for row in rows:
        if col_idx >= len(row["cells"]):
            continue
        cell = row["cells"][col_idx]
        if not is_empty_value(cell.get("value")):
            return True
        merge = cell.get("merge")
        if merge and merge.get("role") == "master":
            return True
    return False


def trim_trailing_empty_columns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return rows

    max_cols = max(len(row["cells"]) for row in rows)
    last_content_col = -1
    for col_idx in range(max_cols):
        if _col_has_content_or_master(rows, col_idx):
            last_content_col = col_idx

    if last_content_col == -1:
        return []

    trimmed_rows: list[dict[str, Any]] = []
    for row in rows:
        cells = row["cells"][: last_content_col + 1]
        trimmed_rows.append({**row, "cells": cells})
    return trimmed_rows


def drop_fully_empty_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        row
        for row in rows
        if row_has_content([cell.get("value") for cell in row["cells"]])
    ]


def build_merge_lookup(worksheet: Worksheet) -> dict[str, tuple[str, str]]:
    """Map every cell address to (merge_range_str, master_address).

    Enables O(1) merge lookup per cell during read-range iteration.
    Both master and slave cells are keyed; master maps to itself.
    """
    lookup: dict[str, tuple[str, str]] = {}
    for merge in worksheet.merged_cells.ranges:
        master_addr = f"{get_column_letter(merge.min_col)}{merge.min_row}"
        range_str = str(merge)
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                addr = f"{get_column_letter(col)}{row}"
                lookup[addr] = (range_str, master_addr)
    return lookup


def build_clean_row(
    worksheet: Worksheet,
    row_index: int,
    from_col_idx: int,
    to_col_idx: int,
    merge_lookup: dict[str, tuple[str, str]] | None = None,
) -> dict[str, Any]:
    cells: list[dict[str, Any]] = []
    for col_idx in range(from_col_idx, to_col_idx + 1):
        column = index_to_column(col_idx)
        address = f"{column}{row_index}"
        value = normalize_scalar(worksheet.cell(row=row_index, column=col_idx).value)
        cell_payload: dict[str, Any] = {
            "column": column,
            "address": address,
            "value": value,
        }
        if merge_lookup and address in merge_lookup:
            range_str, master = merge_lookup[address]
            role = "master" if address == master else "slave"
            cell_payload["merge"] = {
                "range": range_str,
                "role": role,
                "master": master,
            }
        cells.append(cell_payload)
    return {"row": row_index - 1, "cells": cells}
