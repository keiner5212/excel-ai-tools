from __future__ import annotations

import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

from spreadsheet_tools.reader import _serialize_value
from spreadsheet_tools.utils import get_sheet, open_workbook


def _snapshot_dir(abs_file: Path, sheet_name: str) -> Path:
    """Return (and create) the snapshot directory for a file + sheet pair."""
    file_hash = hashlib.sha1(str(abs_file).encode()).hexdigest()[:16]
    safe_sheet = "".join(
        c if (c.isalnum() or c in "_- ") else "_" for c in sheet_name
    )[:50]
    base = (
        Path.home()
        / ".cache"
        / "spreadsheet-tools"
        / "snapshots"
        / file_hash
        / safe_sheet
    )
    base.mkdir(parents=True, exist_ok=True)
    return base


def _snapshot_path(abs_file: Path, sheet_name: str, tag: str) -> Path:
    safe_tag = "".join(c if (c.isalnum() or c in "_-") else "_" for c in tag)
    return _snapshot_dir(abs_file, sheet_name) / f"{safe_tag}.json"


def _resolve_sheet_name(path: str, sheet_name: str | None) -> str:
    workbook = open_workbook(path, read_only=True)
    try:
        return get_sheet(workbook, sheet_name).title
    finally:
        workbook.close()


def create_snapshot(
    path: str,
    *,
    sheet_name: str | None = None,
    tag: str,
    description: str | None = None,
) -> dict[str, Any]:
    """Capture all non-empty cell values and persist to a JSON snapshot file."""
    abs_file = Path(path).expanduser().resolve()
    workbook = open_workbook(path, read_only=False, data_only=True)
    try:
        sheet = get_sheet(workbook, sheet_name)
        actual_sheet = sheet.title

        rows_data: dict[str, dict[str, Any]] = {}
        for row in sheet.iter_rows():
            row_idx = str(row[0].row)
            row_vals: dict[str, Any] = {}
            for cell in row:
                val = _serialize_value(cell.value)
                col_idx = cell.column
                if val is not None and col_idx is not None:
                    row_vals[get_column_letter(col_idx)] = val
            if row_vals:
                rows_data[row_idx] = row_vals

        snapshot: dict[str, Any] = {
            "file": str(abs_file),
            "sheet": actual_sheet,
            "tag": tag,
            "description": description,
            "created_at": datetime.now(tz=timezone.utc).isoformat(),
            "row_count": len(rows_data),
            "rows": rows_data,
        }

        snap_path = _snapshot_path(abs_file, actual_sheet, tag)
        snap_path.write_text(
            json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        return {
            "file": path,
            "sheet": actual_sheet,
            "tag": tag,
            "description": description,
            "created_at": snapshot["created_at"],
            "row_count": len(rows_data),
            "snapshot_path": str(snap_path),
        }
    finally:
        workbook.close()


def diff_snapshots(
    path: str,
    *,
    sheet_name: str | None = None,
    tag_a: str,
    tag_b: str,
) -> dict[str, Any]:
    """Compare two snapshots and return per-cell changes."""
    abs_file = Path(path).expanduser().resolve()
    actual_sheet = _resolve_sheet_name(path, sheet_name)

    path_a = _snapshot_path(abs_file, actual_sheet, tag_a)
    path_b = _snapshot_path(abs_file, actual_sheet, tag_b)

    if not path_a.exists():
        raise FileNotFoundError(
            f"Snapshot {tag_a!r} not found for sheet {actual_sheet!r}"
        )
    if not path_b.exists():
        raise FileNotFoundError(
            f"Snapshot {tag_b!r} not found for sheet {actual_sheet!r}"
        )

    snap_a: dict[str, Any] = json.loads(path_a.read_text(encoding="utf-8"))
    snap_b: dict[str, Any] = json.loads(path_b.read_text(encoding="utf-8"))

    rows_a: dict[str, dict[str, Any]] = snap_a["rows"]
    rows_b: dict[str, dict[str, Any]] = snap_b["rows"]

    changes: list[dict[str, Any]] = []
    added_rows: list[str] = []
    removed_rows: list[str] = []

    all_rows = sorted(set(rows_a.keys()) | set(rows_b.keys()), key=int)

    for row_idx in all_rows:
        in_a = row_idx in rows_a
        in_b = row_idx in rows_b

        if in_a and not in_b:
            removed_rows.append(row_idx)
            continue
        if in_b and not in_a:
            added_rows.append(row_idx)
            continue

        # Both snapshots have data for this row — diff column by column
        cols_a = rows_a[row_idx]
        cols_b = rows_b[row_idx]
        for col in sorted(set(cols_a.keys()) | set(cols_b.keys())):
            val_a = cols_a.get(col)
            val_b = cols_b.get(col)
            if val_a != val_b:
                changes.append(
                    {
                        "address": f"{col}{row_idx}",
                        "before": val_a,
                        "after": val_b,
                    }
                )

    return {
        "file": path,
        "sheet": actual_sheet,
        "tag_a": tag_a,
        "tag_b": tag_b,
        "created_a": snap_a.get("created_at"),
        "created_b": snap_b.get("created_at"),
        "changes": changes,
        "added_rows": added_rows,
        "removed_rows": removed_rows,
        "total_changes": len(changes),
    }


def list_snapshots(
    path: str,
    *,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """List available snapshot tags for a file + sheet."""
    abs_file = Path(path).expanduser().resolve()
    actual_sheet = _resolve_sheet_name(path, sheet_name)

    snap_dir = _snapshot_dir(abs_file, actual_sheet)
    snapshots: list[dict[str, Any]] = []

    for snap_file in sorted(snap_dir.glob("*.json")):
        try:
            data: dict[str, Any] = json.loads(snap_file.read_text(encoding="utf-8"))
            snapshots.append(
                {
                    "tag": data.get("tag", snap_file.stem),
                    "description": data.get("description"),
                    "created_at": data.get("created_at"),
                    "row_count": data.get("row_count"),
                }
            )
        except Exception:
            # Skip snapshots that cannot be parsed
            pass

    return {
        "file": path,
        "sheet": actual_sheet,
        "count": len(snapshots),
        "snapshots": snapshots,
    }
