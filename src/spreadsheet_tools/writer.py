from __future__ import annotations

import re
from typing import Any

from openpyxl import Workbook

from spreadsheet_tools.cleaner import build_merge_lookup
from spreadsheet_tools.styles import apply_style_updates, get_cell_style
from spreadsheet_tools.utils import (
    column_to_index,
    get_sheet,
    index_to_column,
    open_workbook,
    open_workbook_for_write,
    parse_cell_address,
    resolve_output_path,
    safe_save_workbook,
    validate_cell_address,
    validate_cell_range,
    validate_column,
)


def _serialize_value(value: object | None) -> object | None:
    from datetime import date, datetime, time

    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _resolve_write_address(sheet: Any, address: str) -> tuple[str, str | None]:
    """Return (effective_address, warning).

    If address is a slave cell of a merge, redirect to master and warn.
    Writing to slave cells is silently ignored by Excel; only master shows data.
    """
    merge_lookup = build_merge_lookup(sheet)
    if address not in merge_lookup:
        return address, None
    range_str, master = merge_lookup[address]
    if address == master:
        return address, None
    warning = (
        f"Address {address} is a slave cell of merged range {range_str}. "
        f"Redirected to master cell {master}."
    )
    return master, warning


def edit_cell(
    path: str,
    *,
    sheet_name: str | None,
    address: str,
    value: object | None = None,
    clear_value: bool = False,
    style: dict[str, Any] | None = None,
    save: bool = True,
) -> dict[str, Any]:
    normalized_address = validate_cell_address(address)
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        effective_address, warning = _resolve_write_address(sheet, normalized_address)
        cell = sheet[effective_address]

        if clear_value:
            cell.value = None
        elif value is not None:
            cell.value = value

        if style:
            apply_style_updates(cell, style)

        result: dict[str, Any] = {
            "file": path,
            "sheet": sheet.title,
            "address": effective_address,
            "updated": {
                "value": cell.value,
                "style_changed": bool(style),
            },
        }

        if warning:
            result["warning"] = warning

        if save:
            safe_save_workbook(workbook, path)
            result["saved"] = True
        else:
            result["saved"] = False

        return result
    finally:
        workbook.close()


def get_cell_style_info(
    path: str, *, sheet_name: str | None, address: str
) -> dict[str, Any]:
    normalized_address = validate_cell_address(address)
    # data_only=False needed to read number formats and formula-based cells correctly.
    # read_only=False required because read-only mode doesn't expose full cell style objects.
    workbook = open_workbook(path, read_only=False, data_only=False)
    try:
        sheet = get_sheet(workbook, sheet_name)
        style = get_cell_style(sheet, normalized_address)
        if "value" in style:
            style["value"] = _serialize_value(style["value"])
        return {"file": path, "sheet": sheet.title, **style}
    finally:
        workbook.close()


def copy_sheet_structure(
    path: str,
    *,
    source_sheet: str,
    target_sheet: str,
    overwrite: bool = False,
) -> dict[str, Any]:
    workbook = open_workbook_for_write(path)
    try:
        if source_sheet not in workbook.sheetnames:
            raise ValueError(f"Source sheet {source_sheet!r} not found")

        if target_sheet in workbook.sheetnames:
            if not overwrite:
                raise ValueError(f"Target sheet {target_sheet!r} already exists")
            del workbook[target_sheet]

        source = workbook[source_sheet]
        target = workbook.copy_worksheet(source)
        target.title = target_sheet
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "source_sheet": source_sheet,
            "target_sheet": target_sheet,
            "saved": True,
        }
    finally:
        workbook.close()


def batch_edit(
    path: str,
    *,
    sheet_name: str | None = None,
    edits: list[dict[str, Any]],
    dry_run: bool = False,
) -> dict[str, Any]:
    """Apply multiple cell edits atomically in one workbook load/save cycle.

    Each edit dict must contain ``cell`` (address string) and at least one of:
    ``value``, ``clear`` (bool), or ``style`` (dict).

    All addresses are validated before any edit is applied.  If validation
    fails the workbook is never opened for writing.
    """
    if not edits:
        raise ValueError("edits list is empty")

    # Validate all addresses upfront so we fail early before touching the file
    for i, edit in enumerate(edits):
        if "cell" not in edit:
            raise ValueError(f"edits[{i}] missing required 'cell' field")
        validate_cell_address(edit["cell"])

    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        merge_lookup = build_merge_lookup(sheet)
        applied: list[dict[str, Any]] = []

        for edit in edits:
            address = validate_cell_address(edit["cell"])
            effective_addr = address
            warning: str | None = None

            if address in merge_lookup:
                range_str, master = merge_lookup[address]
                if address != master:
                    warning = (
                        f"Address {address} is a slave of merged range {range_str}. "
                        f"Redirected to master cell {master}."
                    )
                    effective_addr = master

            cell = sheet[effective_addr]
            value = edit.get("value")
            clear = edit.get("clear", False)
            style = edit.get("style")

            if clear:
                cell.value = None
            elif value is not None:
                cell.value = value

            if style:
                apply_style_updates(cell, style)

            entry: dict[str, Any] = {
                "cell": edit["cell"],
                "effective_address": effective_addr,
                "new_value": _serialize_value(cell.value),
                "style_changed": bool(style),
            }
            if warning:
                entry["warning"] = warning
            applied.append(entry)

        result: dict[str, Any] = {
            "file": path,
            "sheet": sheet.title,
            "applied": len(applied),
            "dry_run": dry_run,
            "cells": applied,
        }

        if not dry_run:
            safe_save_workbook(workbook, path)
            result["saved"] = True
        else:
            result["saved"] = False

        return result
    finally:
        workbook.close()


def create_empty_workbook(path: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Create new empty workbook with one sheet, overwriting existing file.

    Bare filenames (no directory component) are automatically placed in workspace/.
    """
    resolved = resolve_output_path(path)
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name
    workbook.save(resolved)
    return {"file": str(resolved), "sheet": sheet_name, "created": True}


def merge_cells(
    path: str,
    *,
    sheet_name: str | None,
    cell_range: str,
) -> dict[str, Any]:
    """Merge cells in a rectangular range.

    The master cell (top-left) keeps its value and style. Slave cells are
    cleared by Excel on open. Safe: uses ZIP-merge save to preserve drawings/images.
    """
    top_left, _ = validate_cell_range(cell_range)
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        sheet.merge_cells(cell_range.upper())
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "merged": cell_range.upper(),
            "master": top_left,
            "saved": True,
        }
    finally:
        workbook.close()


def unmerge_cells(
    path: str,
    *,
    sheet_name: str | None,
    cell_range: str,
) -> dict[str, Any]:
    """Remove a merge from a cell range.

    Individual cells regain independent values (previously all null except master).
    Safe: uses ZIP-merge save.
    """
    validate_cell_range(cell_range)
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        sheet.unmerge_cells(cell_range.upper())
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "unmerged": cell_range.upper(),
            "saved": True,
        }
    finally:
        workbook.close()


def set_column_width(
    path: str,
    *,
    sheet_name: str | None,
    col: str,
    width: float,
) -> dict[str, Any]:
    """Set width of a single column. Width is in Excel character units."""
    normalized_col = validate_column(col)
    if width <= 0:
        raise ValueError(f"Column width must be > 0, got {width}")
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        sheet.column_dimensions[normalized_col].width = width
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "column": normalized_col,
            "width": width,
            "saved": True,
        }
    finally:
        workbook.close()


def set_row_height(
    path: str,
    *,
    sheet_name: str | None,
    row: int,
    height: float,
) -> dict[str, Any]:
    """Set height of a single row. Row is 1-based (Excel row number). Height in points."""
    if row < 1:
        raise ValueError(f"Row number must be >= 1, got {row}")
    if height <= 0:
        raise ValueError(f"Row height must be > 0, got {height}")
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        sheet.row_dimensions[row].height = height
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "row": row,
            "height": height,
            "saved": True,
        }
    finally:
        workbook.close()


def batch_set_dimensions(
    path: str,
    *,
    sheet_name: str | None = None,
    column_widths: list[dict[str, Any]] | None = None,
    row_heights: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Set multiple column widths and/or row heights in a single save.

    column_widths: list of {"col": "A", "width": 15}
    row_heights:   list of {"row": 1, "height": 30}  (row is 1-based)
    """
    if not column_widths and not row_heights:
        raise ValueError("At least one of column_widths or row_heights must be provided")

    # Validate all inputs before touching the file
    validated_cols: list[tuple[str, float]] = []
    for item in column_widths or []:
        col = validate_column(item["col"])
        w = float(item["width"])
        if w <= 0:
            raise ValueError(f"Column width must be > 0, got {w} for col {col}")
        validated_cols.append((col, w))

    validated_rows: list[tuple[int, float]] = []
    for item in row_heights or []:
        row_num = int(item["row"])
        h = float(item["height"])
        if row_num < 1:
            raise ValueError(f"Row number must be >= 1, got {row_num}")
        if h <= 0:
            raise ValueError(f"Row height must be > 0, got {h}")
        validated_rows.append((row_num, h))

    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        for col, width in validated_cols:
            sheet.column_dimensions[col].width = width
        for row_num, height in validated_rows:
            sheet.row_dimensions[row_num].height = height
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "column_widths": [{"col": c, "width": w} for c, w in validated_cols],
            "row_heights": [{"row": r, "height": h} for r, h in validated_rows],
            "saved": True,
        }
    finally:
        workbook.close()


def freeze_panes(
    path: str,
    *,
    sheet_name: str | None,
    cell: str,
) -> dict[str, Any]:
    """Freeze rows above and columns left of the given cell.

    Examples: cell=B2 freezes row 1 and column A.
              cell=A2 freezes only row 1.
              cell=B1 freezes only column A.
    """
    normalized_cell = validate_cell_address(cell)
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        sheet.freeze_panes = normalized_cell
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "freeze_panes": normalized_cell,
            "saved": True,
        }
    finally:
        workbook.close()


def unfreeze_panes(
    path: str,
    *,
    sheet_name: str | None,
) -> dict[str, Any]:
    """Remove all freeze panes from a sheet."""
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        sheet.freeze_panes = None
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "freeze_panes": None,
            "saved": True,
        }
    finally:
        workbook.close()


def add_sheet(
    path: str,
    *,
    sheet_name: str,
    position: int | None = None,
) -> dict[str, Any]:
    """Add a new empty sheet to an existing workbook."""
    workbook = open_workbook_for_write(path)
    try:
        if sheet_name in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} already exists")
        ws = workbook.create_sheet(title=sheet_name, index=position)
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": ws.title,
            "index": workbook.sheetnames.index(ws.title),
            "added": True,
            "saved": True,
        }
    finally:
        workbook.close()


def rename_sheet(
    path: str,
    *,
    old_name: str,
    new_name: str,
) -> dict[str, Any]:
    """Rename a sheet. Fails if old_name not found or new_name already taken."""
    workbook = open_workbook_for_write(path)
    try:
        if old_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"Sheet {old_name!r} not found. Available: {available}"
            )
        if new_name in workbook.sheetnames:
            raise ValueError(f"Sheet {new_name!r} already exists")
        workbook[old_name].title = new_name
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "old_name": old_name,
            "new_name": new_name,
            "saved": True,
        }
    finally:
        workbook.close()


def format_range(
    path: str,
    *,
    sheet_name: str | None,
    cell_range: str,
    style: dict[str, Any],
) -> dict[str, Any]:
    """Apply the same style to every master cell in a rectangular range.

    Slave cells (merged non-master cells) are skipped silently. Useful for
    applying header colors, fonts, or number formats to entire rows/columns.
    """
    top_left, bottom_right = validate_cell_range(cell_range)
    tl_col, tl_row = parse_cell_address(top_left)
    br_col, br_row = parse_cell_address(bottom_right)
    from_col_idx = column_to_index(tl_col)
    to_col_idx = column_to_index(br_col)

    if from_col_idx > to_col_idx or tl_row > br_row:
        raise ValueError(
            f"Range {cell_range!r}: top-left must be above and left of bottom-right"
        )

    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        merge_lookup = build_merge_lookup(sheet)
        affected: list[str] = []

        for row_num in range(tl_row, br_row + 1):
            for col_idx in range(from_col_idx, to_col_idx + 1):
                address = f"{index_to_column(col_idx)}{row_num}"
                # Skip slave cells — only style master cells to avoid ghost styles
                if address in merge_lookup:
                    _, master = merge_lookup[address]
                    if address != master:
                        continue
                apply_style_updates(sheet[address], style)
                affected.append(address)

        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "range": cell_range.upper(),
            "cells_styled": len(affected),
            "saved": True,
        }
    finally:
        workbook.close()


def set_tab_color(
    path: str,
    *,
    sheet_name: str | None,
    color: str,
) -> dict[str, Any]:
    """Set the sheet tab color. Color is a 6-char hex RGB string, e.g. '1F4E79'."""
    normalized = color.strip().upper().lstrip("#")
    if len(normalized) not in (6, 8):
        raise ValueError(
            f"Color must be a 6-char RGB hex (e.g. '1F4E79'), got {color!r}"
        )
    # Ensure ARGB format: prefix FF if only 6 chars
    argb = normalized if len(normalized) == 8 else f"FF{normalized}"
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        sheet.sheet_properties.tabColor = argb
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "tab_color": argb,
            "saved": True,
        }
    finally:
        workbook.close()


def toggle_auto_filter(
    path: str,
    *,
    sheet_name: str | None,
    cell_range: str | None,
) -> dict[str, Any]:
    """Enable auto-filter on a range, or clear it if cell_range is None.

    cell_range: e.g. 'A8:E8' — sets the auto-filter header row.
    Pass cell_range=None to remove existing auto-filter.
    """
    if cell_range is not None:
        validate_cell_range(cell_range)
    workbook = open_workbook_for_write(path)
    try:
        sheet = get_sheet(workbook, sheet_name)
        if cell_range is not None:
            sheet.auto_filter.ref = cell_range.upper()
        else:
            sheet.auto_filter.ref = None
        safe_save_workbook(workbook, path)
        return {
            "file": path,
            "sheet": sheet.title,
            "auto_filter": cell_range.upper() if cell_range else None,
            "saved": True,
        }
    finally:
        workbook.close()


def find_replace(
    path: str,
    *,
    query: str,
    replace_with: Any = None,
    sheet_name: str | None = None,
    case_sensitive: bool = False,
    use_regex: bool = False,
    dry_run: bool = False,
    max_results: int = 200,
) -> dict[str, Any]:
    """Search cell values and optionally replace matches.

    Default matching is exact-value: the entire string representation of the
    cell value must equal the query.  Pass ``use_regex=True`` for
    substring/pattern matching (``re.search``); replacement is then applied
    via ``re.sub`` on the string representation.

    ``replace_with`` accepts any Python value (int, float, str).  Numeric
    values are stored as numbers in Excel, not text.  When ``use_regex=True``
    the replacement is forced to str (required by ``re.sub``).

    Two-pass approach: read pass collects matches with data_only=True (computed
    values); write pass applies replacements using a separate workbook load.
    """
    from spreadsheet_tools.reader import _serialize_value

    flags = 0 if case_sensitive else re.IGNORECASE
    pattern: re.Pattern[str] | None = re.compile(query, flags) if use_regex else None
    needle = query if case_sensitive else query.lower()

    # --- Pass 1: find matches (read-only, computed values) ---
    matches: list[dict[str, Any]] = []
    read_wb = open_workbook(path, read_only=False, data_only=True)
    try:
        target_sheets = (
            [get_sheet(read_wb, sheet_name)]
            if sheet_name
            else [read_wb[name] for name in read_wb.sheetnames]
        )
        for sheet in target_sheets:
            if len(matches) >= max_results:
                break
            for row in sheet.iter_rows():
                if len(matches) >= max_results:
                    break
                for cell in row:
                    if cell.value is None:
                        continue
                    val_str = str(cell.value)
                    if use_regex:
                        matched = bool(pattern.search(val_str))  # type: ignore[union-attr]
                    else:
                        compare = val_str if case_sensitive else val_str.lower()
                        matched = needle == compare
                    if matched:
                        matches.append(
                            {
                                "sheet": sheet.title,
                                "address": cell.coordinate,
                                "value": _serialize_value(cell.value),
                            }
                        )
    finally:
        read_wb.close()

    result: dict[str, Any] = {
        "file": path,
        "query": query,
        "total": len(matches),
        "truncated": len(matches) >= max_results,
        "matches": matches,
    }

    if replace_with is None:
        return result

    result["replace_with"] = replace_with

    # Compute new values for each match
    replacements: list[dict[str, Any]] = []
    for m in matches:
        if use_regex:
            # re.sub requires string operands; result is always str
            new_val: Any = re.sub(
                query, str(replace_with), str(m["value"]), flags=flags
            )
        else:
            # Preserve the caller's type (int, float, str, …)
            new_val = replace_with
        replacements.append(
            {
                "sheet": m["sheet"],
                "address": m["address"],
                "before": m["value"],
                "after": new_val,
            }
        )

    if dry_run:
        result["dry_run"] = True
        result["replacements_preview"] = replacements
        result["saved"] = False
        return result

    # --- Pass 2: apply replacements ---
    write_wb = open_workbook_for_write(path)
    try:
        for rep in replacements:
            ws = get_sheet(write_wb, rep["sheet"])
            ws[rep["address"]].value = rep["after"]  # type: ignore[union-attr]
        safe_save_workbook(write_wb, path)
    finally:
        write_wb.close()

    result["dry_run"] = False
    result["replacements"] = replacements
    result["saved"] = True
    return result
