from __future__ import annotations

from datetime import date, datetime, time
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_tools.cleaner import (
    build_clean_row,
    build_merge_lookup,
    drop_fully_empty_rows,
    trim_trailing_empty_columns,
    trim_trailing_empty_rows,
)
from spreadsheet_tools.utils import (
    GENERIC_NAME_RE,
    SECTION_HEADER_RE,
    column_to_index,
    get_sheet,
    index_to_column,
    keyword_overlap_ratio,
    normalize_scalar,
    open_workbook,
)


def _serialize_value(value: object | None) -> object | None:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return normalize_scalar(value)


def list_sheets(path: str) -> dict[str, Any]:
    workbook = open_workbook(path, read_only=True)
    try:
        active = workbook.active.title if workbook.active else None
        sheets = []
        for idx, name in enumerate(workbook.sheetnames):
            sheet = workbook[name]
            sheets.append(
                {
                    "name": name,
                    "index": idx,
                    "state": sheet.sheet_state,
                }
            )
        return {
            "file": path,
            "active_sheet": active,
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    finally:
        workbook.close()


def workbook_info(path: str) -> dict[str, Any]:
    workbook = open_workbook(path, read_only=True)
    try:
        properties = workbook.properties
        return {
            "file": path,
            "format": "xlsm" if path.lower().endswith(".xlsm") else "xlsx",
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": list(workbook.sheetnames),
            "active_sheet": workbook.active.title if workbook.active else None,
            "properties": {
                "title": properties.title,
                "creator": properties.creator,
                "created": properties.created.isoformat()
                if properties.created
                else None,
                "modified": properties.modified.isoformat()
                if properties.modified
                else None,
                "subject": properties.subject,
                "description": properties.description,
            },
        }
    finally:
        workbook.close()


def sheet_info(path: str, sheet_name: str | None = None) -> dict[str, Any]:
    workbook = open_workbook(path, read_only=False)
    try:
        sheet = get_sheet(workbook, sheet_name)
        merged = [str(item) for item in sheet.merged_cells.ranges]
        return {
            "file": path,
            "sheet": sheet.title,
            "dimensions": {
                "min_row": sheet.min_row or 1,
                "max_row": sheet.max_row or 1,
                "min_col": get_column_letter(sheet.min_column) if sheet.min_column else "A",
                "max_col": get_column_letter(sheet.max_column) if sheet.max_column else "A",
            },
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "max_column_letter": get_column_letter(sheet.max_column)
            if sheet.max_column
            else None,
            "merged_ranges": merged,
            "freeze_panes": sheet.freeze_panes,
            "auto_filter": sheet.auto_filter.ref,
        }
    finally:
        workbook.close()


def read_range(
    path: str,
    *,
    sheet_name: str | None = None,
    from_col: str,
    to_col: str,
    from_row: int,
    to_row: int,
    include_empty_rows: bool = False,
    trim_empty: bool = True,
    include_formulas: bool = False,
) -> dict[str, Any]:
    if from_row < 0 or to_row < 0:
        raise ValueError("Row indices are zero-based and must be >= 0")
    if from_row > to_row:
        raise ValueError("from_row must be <= to_row")

    from_col_idx = column_to_index(from_col)
    to_col_idx = column_to_index(to_col)
    if from_col_idx > to_col_idx:
        raise ValueError("from_col must be <= to_col")

    data_only = not include_formulas
    workbook = open_workbook(path, read_only=False, data_only=data_only)
    try:
        sheet = get_sheet(workbook, sheet_name)
        merge_lookup = build_merge_lookup(sheet)
        rows: list[dict[str, Any]] = []
        for excel_row in range(from_row + 1, to_row + 2):
            row_payload = build_clean_row(
                sheet, excel_row, from_col_idx, to_col_idx, merge_lookup
            )
            for cell in row_payload["cells"]:
                cell["value"] = _serialize_value(cell["value"])
            rows.append(row_payload)

        if trim_empty:
            rows = trim_trailing_empty_columns(rows)
            if not include_empty_rows:
                rows = drop_fully_empty_rows(rows)
                rows = trim_trailing_empty_rows(rows)

        merged_in_range = _merged_ranges_in_area(
            sheet,
            from_col_idx,
            to_col_idx,
            from_row + 1,
            to_row + 1,
        )

        return {
            "file": path,
            "sheet": sheet.title,
            "range": {
                "from_col": index_to_column(from_col_idx),
                "to_col": index_to_column(to_col_idx),
                "from_row": from_row,
                "to_row": to_row,
            },
            "row_count": len(rows),
            "rows": rows,
            "merged_ranges": merged_in_range,
            "notes": [
                "Values only; formatting metadata excluded.",
                "Rows are zero-based in output.",
                "Empty rows/columns trimmed unless include_empty_rows is enabled.",
                "Merged cells: 'master' is the top-left cell of the range (write target). 'slave' cells are non-writable; their value is always null.",
            ],
        }
    finally:
        workbook.close()


def _merged_ranges_in_area(
    sheet: Worksheet,
    from_col_idx: int,
    to_col_idx: int,
    from_row: int,
    to_row: int,
) -> list[str]:
    merged: list[str] = []
    for item in sheet.merged_cells.ranges:
        if (
            item.min_col <= to_col_idx
            and item.max_col >= from_col_idx
            and item.min_row <= to_row
            and item.max_row >= from_row
        ):
            merged.append(str(item))
    return merged


def find_values(
    path: str,
    *,
    query: str,
    sheet_name: str | None = None,
    case_sensitive: bool = False,
    max_results: int = 50,
) -> dict[str, Any]:
    workbook = open_workbook(path, read_only=True)
    try:
        sheets = (
            [get_sheet(workbook, sheet_name)]
            if sheet_name
            else [workbook[name] for name in workbook.sheetnames]
        )
        needle = query if case_sensitive else query.casefold()
        matches: list[dict[str, Any]] = []

        for sheet in sheets:
            # read_only mode doesn't support merged_cells; skip merge lookup there
            try:
                merge_lookup = build_merge_lookup(sheet)  # type: ignore[arg-type]
            except AttributeError:
                merge_lookup = {}

            for row in sheet.iter_rows():
                for cell in row:
                    value = _serialize_value(cell.value)
                    if value is None:
                        continue
                    haystack = str(value)
                    compare = haystack if case_sensitive else haystack.casefold()
                    if needle in compare:
                        if len(matches) >= max_results:
                            return {
                                "file": path,
                                "query": query,
                                "match_count": len(matches),
                                "truncated": True,
                                "matches": matches,
                            }
                        match: dict[str, Any] = {
                            "sheet": sheet.title,
                            "address": cell.coordinate,
                            "value": value,
                        }
                        if cell.coordinate in merge_lookup:
                            range_str, master = merge_lookup[cell.coordinate]
                            role = "master" if cell.coordinate == master else "slave"
                            match["merge"] = {
                                "range": range_str,
                                "role": role,
                                "master": master,
                            }
                        matches.append(match)

        return {
            "file": path,
            "query": query,
            "match_count": len(matches),
            "truncated": False,
            "matches": matches,
        }
    finally:
        workbook.close()


def read_cell(
    path: str, *, sheet_name: str | None, address: str, include_formulas: bool = False
) -> dict[str, Any]:
    # read_only=True uses ReadOnlyWorksheet which doesn't support sheet[addr] access.
    workbook = open_workbook(path, read_only=False, data_only=not include_formulas)
    try:
        sheet = get_sheet(workbook, sheet_name)
        cell = sheet[address.upper()]
        return {
            "file": path,
            "sheet": sheet.title,
            "address": cell.coordinate,
            "value": _serialize_value(cell.value),
        }
    finally:
        workbook.close()


def section_map(
    path: str,
    *,
    sheet_name: str | None = None,
    min_row: int = 0,
    max_row: int | None = None,
) -> dict[str, Any]:
    """Discover numbered section headers and map each to its row range.

    Scans columns B then A for cells matching the pattern ``N.N.N Title``.
    Each section's row range ends one row before the next section's header.
    """
    workbook = open_workbook(path, read_only=False, data_only=True)
    try:
        sheet = get_sheet(workbook, sheet_name)
        effective_max = (
            max_row if max_row is not None else max((sheet.max_row or 1) - 1, 0)
        )

        sections: list[dict[str, Any]] = []

        for excel_row in range(min_row + 1, effective_max + 2):
            # Prefer column B (index 2); fall back to A (index 1)
            for col_idx in (2, 1):
                raw = sheet.cell(row=excel_row, column=col_idx).value
                if not (raw and isinstance(raw, str)):
                    continue
                text = raw.strip()
                m = SECTION_HEADER_RE.match(text)
                if not m:
                    continue
                prefix = m.group(1).rstrip(".")
                title = m.group(2).strip()
                depth = len(prefix.split("."))
                sections.append(
                    {
                        "header": text,
                        "prefix": prefix,
                        "title": title,
                        "depth": depth,
                        "header_row": excel_row - 1,  # zero-based
                        "column": index_to_column(col_idx),
                        "row_range": {"from": excel_row - 1, "to": None},
                    }
                )
                break  # matched in col B; no need to check col A

        # Assign end rows: each section ends one row before the next header
        for i, sec in enumerate(sections):
            next_start = (
                sections[i + 1]["header_row"]
                if i < len(sections) - 1
                else effective_max
            )
            sec["row_range"]["to"] = (
                next_start - 1 if i < len(sections) - 1 else effective_max
            )

        return {
            "file": path,
            "sheet": sheet.title,
            "scanned_rows": {"min": min_row, "max": effective_max},
            "section_count": len(sections),
            "sections": sections,
        }
    finally:
        workbook.close()


def audit_range(
    path: str,
    *,
    sheet_name: str | None = None,
    from_col: str,
    to_col: str,
    from_row: int,
    to_row: int,
    show_slaves: bool = False,
) -> dict[str, Any]:
    """Audit every master cell in a range, flagging empty ones.

    Only master cells (or unmerged cells) are reported by default.
    Pass show_slaves=True to include slave cells in output.
    """
    if from_row < 0 or to_row < 0:
        raise ValueError("Row indices are zero-based and must be >= 0")
    if from_row > to_row:
        raise ValueError("from_row must be <= to_row")

    from_col_idx = column_to_index(from_col)
    to_col_idx = column_to_index(to_col)
    if from_col_idx > to_col_idx:
        raise ValueError("from_col must be <= to_col")

    workbook = open_workbook(path, read_only=False, data_only=True)
    try:
        sheet = get_sheet(workbook, sheet_name)
        merge_lookup = build_merge_lookup(sheet)

        audit_rows: list[dict[str, Any]] = []
        total_master = 0
        empty_master = 0

        for excel_row in range(from_row + 1, to_row + 2):
            row_cells: list[dict[str, Any]] = []

            for col_idx in range(from_col_idx, to_col_idx + 1):
                column = index_to_column(col_idx)
                address = f"{column}{excel_row}"
                value = _serialize_value(
                    sheet.cell(row=excel_row, column=col_idx).value
                )

                is_master = True
                merge_info: dict[str, Any] | None = None

                if address in merge_lookup:
                    range_str, master = merge_lookup[address]
                    role = "master" if address == master else "slave"
                    is_master = role == "master"
                    merge_info = {"range": range_str, "role": role, "master": master}

                is_empty = value is None or (
                    isinstance(value, str) and not value.strip()
                )

                if is_master:
                    total_master += 1
                    if is_empty:
                        empty_master += 1

                if not (is_master or show_slaves):
                    continue

                cell_entry: dict[str, Any] = {
                    "address": address,
                    "value": value,
                    "is_master": is_master,
                    "is_empty": is_empty,
                }
                if merge_info:
                    cell_entry["merge"] = merge_info
                row_cells.append(cell_entry)

            if row_cells:
                audit_rows.append({"row": excel_row - 1, "cells": row_cells})

        return {
            "file": path,
            "sheet": sheet.title,
            "range": {
                "from_col": from_col.upper(),
                "to_col": to_col.upper(),
                "from_row": from_row,
                "to_row": to_row,
            },
            "audit": audit_rows,
            "summary": {
                "total_rows": to_row - from_row + 1,
                "total_master_cells": total_master,
                "empty_master_cells": empty_master,
                "filled_master_cells": total_master - empty_master,
            },
        }
    finally:
        workbook.close()


def describe_section(
    path: str,
    *,
    sheet_name: str | None = None,
    from_data_row: int,
    to_data_row: int,
    name_col: str = "B",
    desc_col: str = "D",
    cost_col: str | None = "L",
    header_row: int | None = None,
) -> dict[str, Any]:
    """Audit a strategy-table section for name/description consistency.

    Checks each data row for:
    - Generic placeholder names matching ``Estrategia de X N`` pattern.
    - Empty description when name is present.
    - Keyword mismatch between name and description (overlap < 30 %).
    - Empty cost cell (warning only).
    """
    if from_data_row < 0 or to_data_row < 0:
        raise ValueError("Row indices are zero-based and must be >= 0")
    if from_data_row > to_data_row:
        raise ValueError("from_data_row must be <= to_data_row")

    name_col_idx = column_to_index(name_col)
    desc_col_idx = column_to_index(desc_col)
    cost_col_idx = column_to_index(cost_col) if cost_col else None

    workbook = open_workbook(path, read_only=False, data_only=True)
    try:
        sheet = get_sheet(workbook, sheet_name)
        merge_lookup = build_merge_lookup(sheet)

        def _master_value(excel_row: int, col_idx: int) -> tuple[str, Any]:
            col_letter = index_to_column(col_idx)
            addr = f"{col_letter}{excel_row}"
            if addr in merge_lookup:
                _, master = merge_lookup[addr]
                addr = master
            return addr, _serialize_value(sheet[addr].value)

        rows: list[dict[str, Any]] = []
        error_count = 0
        warning_count = 0
        affected_rows: set[int] = set()

        for excel_row in range(from_data_row + 1, to_data_row + 2):
            name_addr, name_val = _master_value(excel_row, name_col_idx)
            desc_addr, desc_val = _master_value(excel_row, desc_col_idx)

            # Skip entirely blank rows
            if name_val is None and desc_val is None:
                continue

            cost_entry: dict[str, Any] | None = None
            if cost_col_idx is not None:
                cost_addr, cost_val = _master_value(excel_row, cost_col_idx)
                cost_entry = {"address": cost_addr, "value": cost_val}

            issues: list[dict[str, Any]] = []

            if (
                name_val
                and isinstance(name_val, str)
                and GENERIC_NAME_RE.match(name_val)
            ):
                issues.append(
                    {
                        "severity": "warning",
                        "type": "generic_name",
                        "detail": (
                            f"{name_addr}: {name_val!r} matches "
                            f"generic 'Estrategia de X N' pattern"
                        ),
                    }
                )
                warning_count += 1

            if name_val and not desc_val:
                issues.append(
                    {
                        "severity": "error",
                        "type": "missing_description",
                        "detail": f"{desc_addr}: description is empty for name {str(name_val)!r}",
                    }
                )
                error_count += 1

            if (
                name_val
                and desc_val
                and isinstance(name_val, str)
                and isinstance(desc_val, str)
            ):
                ratio = keyword_overlap_ratio(name_val, desc_val)
                if ratio < 0.3:
                    issues.append(
                        {
                            "severity": "error",
                            "type": "name_desc_mismatch",
                            "detail": (
                                f"{name_addr} name {str(name_val)!r} not reflected in "
                                f"{desc_addr} description (overlap {ratio:.0%})"
                            ),
                        }
                    )
                    error_count += 1

            if cost_entry and cost_entry["value"] is None:
                issues.append(
                    {
                        "severity": "warning",
                        "type": "missing_cost",
                        "detail": f"{cost_entry['address']}: cost is empty",
                    }
                )
                warning_count += 1

            if issues:
                affected_rows.add(excel_row - 1)

            row_result: dict[str, Any] = {
                "row": excel_row - 1,
                "name": {"address": name_addr, "value": name_val},
                "description": {"address": desc_addr, "value": desc_val},
                "issues": issues,
            }
            if cost_entry is not None:
                row_result["cost"] = cost_entry
            rows.append(row_result)

        total = len(rows)
        return {
            "file": path,
            "sheet": sheet.title,
            "section": {
                "header_row": header_row,
                "data_rows": f"{from_data_row}-{to_data_row}",
                "name_col": name_col.upper(),
                "desc_col": desc_col.upper(),
                "cost_col": cost_col.upper() if cost_col else None,
            },
            "rows": rows,
            "summary": {
                "total": total,
                "errors": error_count,
                "warnings": warning_count,
                "ok": total - len(affected_rows),
            },
        }
    finally:
        workbook.close()


def validate_rules(
    path: str,
    *,
    sheet_name: str | None = None,
    rules: list[str],
) -> dict[str, Any]:
    """Apply a list of validation rules and return pass/fail per rule.

    See ``spreadsheet_tools.rules.apply_rule`` for supported rule syntax.
    """
    if not rules:
        raise ValueError("At least one rule is required")

    from spreadsheet_tools.rules import apply_rule

    workbook = open_workbook(path, read_only=False, data_only=True)
    try:
        sheet = get_sheet(workbook, sheet_name)
        # Build merge lookup once; passed to every rule so slave cells resolve correctly
        merge_lookup = build_merge_lookup(sheet)
        results: list[dict[str, Any]] = []
        for rule_text in rules:
            result = apply_rule(rule_text, sheet, merge_lookup)
            results.append(
                {
                    "rule": result.rule,
                    "passed": result.passed,
                    "message": result.message,
                    "addresses": result.addresses,
                }
            )

        total = len(results)
        passed_count = sum(1 for r in results if r["passed"])
        return {
            "file": path,
            "sheet": sheet.title,
            "results": results,
            "summary": {
                "total": total,
                "passed": passed_count,
                "failed": total - passed_count,
            },
        }
    finally:
        workbook.close()
