from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from spreadsheet_tools.utils import parse_cell_address


def _color_to_str(color: object | None) -> str | None:
    if color is None:
        return None

    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        rgb = getattr(color, "rgb", None)
        return rgb if isinstance(rgb, str) else None
    if color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        return f"indexed:{indexed}" if indexed is not None else None
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        tint = getattr(color, "tint", 0.0)
        return f"theme:{theme}:tint:{tint}" if theme is not None else None
    if color_type == "auto":
        return "auto"

    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and "Values must be of type" not in rgb:
        return rgb
    return None


def _side_to_dict(side: Side | None) -> dict[str, Any] | None:
    if side is None or side.style is None:
        return None
    return {
        "style": side.style,
        "color": _color_to_str(side.color),
    }


def _border_to_dict(border: Border | None) -> dict[str, Any] | None:
    if border is None:
        return None
    payload = {
        "left": _side_to_dict(border.left),
        "right": _side_to_dict(border.right),
        "top": _side_to_dict(border.top),
        "bottom": _side_to_dict(border.bottom),
        "diagonal": _side_to_dict(border.diagonal),
    }
    if all(value is None for value in payload.values()):
        return None
    return payload


def _font_to_dict(font: Font | None) -> dict[str, Any] | None:
    if font is None:
        return None
    payload = {
        "name": font.name,
        "size": font.size,
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strike": font.strike,
        "color": _color_to_str(font.color),
    }
    if payload == {
        "name": "Calibri",
        "size": 11.0,
        "bold": False,
        "italic": False,
        "underline": None,
        "strike": False,
        "color": None,
    }:
        return None
    return payload


def _fill_to_dict(fill: PatternFill | None) -> dict[str, Any] | None:
    if fill is None or fill.fill_type is None:
        return None
    return {
        "fill_type": fill.fill_type,
        "start_color": _color_to_str(fill.start_color),
        "end_color": _color_to_str(fill.end_color),
    }


def _alignment_to_dict(alignment: Alignment | None) -> dict[str, Any] | None:
    if alignment is None:
        return None
    payload = {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "wrap_text": alignment.wrap_text,
        "shrink_to_fit": alignment.shrink_to_fit,
        "indent": alignment.indent,
        "text_rotation": alignment.text_rotation,
    }
    if all(value in (None, False, 0) for value in payload.values()):
        return None
    return payload


def get_cell_style(worksheet: Worksheet, address: str) -> dict[str, Any]:
    column, row = parse_cell_address(address)
    cell = worksheet[f"{column}{row}"]
    number_format = (
        cell.number_format
        if cell.number_format and cell.number_format != "General"
        else None
    )
    style = {
        "address": address.upper(),
        "value": cell.value,
        "data_type": cell.data_type,
        "number_format": number_format,
        "font": _font_to_dict(cell.font),
        "fill": _fill_to_dict(cell.fill),
        "alignment": _alignment_to_dict(cell.alignment),
        "border": _border_to_dict(cell.border),
        "protection": {
            "locked": cell.protection.locked if cell.protection else None,
            "hidden": cell.protection.hidden if cell.protection else None,
        },
        "comment": cell.comment.text if cell.comment else None,
        "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
    }
    return {
        key: value
        for key, value in style.items()
        if value not in (None, {"locked": True, "hidden": False})
    }


def _parse_color(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip().upper()
    if normalized.startswith("#"):
        normalized = normalized[1:]
    if normalized.startswith("FF") and len(normalized) == 8:
        return normalized
    if len(normalized) == 6:
        return f"FF{normalized}"
    return normalized


def apply_style_updates(cell: Any, style: dict[str, Any]) -> None:
    if "font" in style and isinstance(style["font"], dict):
        font_data = style["font"]
        current = cell.font
        cell.font = Font(
            name=font_data.get("name", current.name),
            size=font_data.get("size", current.size),
            bold=font_data.get("bold", current.bold),
            italic=font_data.get("italic", current.italic),
            underline=font_data.get("underline", current.underline),
            strike=font_data.get("strike", current.strike),
            color=_parse_color(font_data.get("color"))
            if font_data.get("color")
            else current.color,
        )

    if "fill" in style and isinstance(style["fill"], dict):
        fill_data = style["fill"]
        cell.fill = PatternFill(
            fill_type=fill_data.get("fill_type", "solid"),
            start_color=_parse_color(fill_data.get("start_color")),
            end_color=_parse_color(fill_data.get("end_color")),
        )

    if "alignment" in style and isinstance(style["alignment"], dict):
        align_data = style["alignment"]
        current = cell.alignment
        cell.alignment = Alignment(
            horizontal=align_data.get("horizontal", current.horizontal),
            vertical=align_data.get("vertical", current.vertical),
            wrap_text=align_data.get("wrap_text", current.wrap_text),
            shrink_to_fit=align_data.get("shrink_to_fit", current.shrink_to_fit),
            indent=align_data.get("indent", current.indent),
            text_rotation=align_data.get("text_rotation", current.text_rotation),
        )

    if "number_format" in style:
        cell.number_format = style["number_format"]

    if "comment" in style:
        from openpyxl.comments import Comment

        comment_text = style["comment"]
        cell.comment = (
            Comment(comment_text, "spreadsheet-tools") if comment_text else None
        )

    if "hyperlink" in style:
        from openpyxl.worksheet.hyperlink import Hyperlink

        target = style["hyperlink"]
        cell.hyperlink = (
            Hyperlink(ref=cell.coordinate, target=target) if target else None
        )
