from __future__ import annotations

import io
import os
import re
import shutil
import tempfile
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

COL_RE = re.compile(r"^[A-Z]+$")
# Row 0 is invalid in Excel (1-based); [1-9]\d* rejects it at validation time.
CELL_RE = re.compile(r"^([A-Z]+)([1-9]\d*)$")

# Register XML namespaces once at import time (module-level side effect is safe;
# repeated calls from multiple threads would corrupt the global namespace registry).
_RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
ET.register_namespace("", _RELS_NS)


def resolve_path(path: str) -> Path:
    resolved = Path(path).expanduser().resolve()
    if not resolved.exists():
        raise FileNotFoundError(f"File not found: {resolved}")
    return resolved


def open_workbook(
    path: str, *, read_only: bool = False, data_only: bool = True
) -> Workbook:
    return load_workbook(
        resolve_path(path), read_only=read_only, data_only=data_only, keep_vba=True
    )


def open_workbook_for_write(path: str) -> Workbook:
    return load_workbook(resolve_path(path), data_only=False, keep_vba=True)


def validate_column(column: str) -> str:
    normalized = column.strip().upper()
    if not COL_RE.match(normalized):
        raise ValueError(f"Invalid column letter: {column!r}")
    return normalized


def validate_cell_address(address: str) -> str:
    normalized = address.strip().upper()
    if not CELL_RE.match(normalized):
        raise ValueError(f"Invalid cell address: {address!r}")
    return normalized


def column_to_index(column: str) -> int:
    return column_index_from_string(validate_column(column))


def index_to_column(index: int) -> str:
    if index < 1:
        raise ValueError(f"Column index must be >= 1, got {index}")
    return get_column_letter(index)


def parse_cell_address(address: str) -> tuple[str, int]:
    match = CELL_RE.match(validate_cell_address(address))
    assert match is not None
    return match.group(1), int(match.group(2))


def get_sheet(workbook: Workbook, sheet_name: str | None) -> Worksheet:
    # Check membership before access to avoid a raw KeyError from workbook[name].
    if sheet_name is not None and sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet {sheet_name!r} not found. Available: {available}")
    sheet = workbook.active if sheet_name is None else workbook[sheet_name]
    if sheet is None:
        raise ValueError("No active sheet found in workbook")
    if not isinstance(sheet, (Worksheet, ReadOnlyWorksheet)):
        raise ValueError(
            f"Sheet {sheet_name!r} is not a worksheet (may be a chart sheet)"
        )
    return sheet  # type: ignore[return-value]


def normalize_scalar(value: object) -> object | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _merge_rels(
    orig_data: bytes, new_data: bytes, *, prefer_new: bool = False
) -> bytes:
    """Merge two .rels XML files.

    Default (prefer_new=False): starts from original, appends any new entries
    from openpyxl that have a different Id. Used for worksheet rels to preserve
    drawing/hyperlink references openpyxl drops.

    prefer_new=True: starts from openpyxl's version, appends any original entries
    not already present. Used for workbook.xml.rels because openpyxl renumbers
    rIds in workbook.xml, so the rels must match openpyxl's numbering scheme.
    """
    tag = f"{{{_RELS_NS}}}Relationship"

    orig_root = ET.fromstring(orig_data)
    new_root = ET.fromstring(new_data)

    if prefer_new:
        base_root = new_root
        extra_root = orig_root
    else:
        base_root = orig_root
        extra_root = new_root

    base_ids = {el.get("Id") for el in base_root.iter(tag)}
    for el in extra_root.iter(tag):
        if el.get("Id") not in base_ids:
            base_root.append(el)

    xml_str = ET.tostring(base_root, encoding="unicode")
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_str
    ).encode("utf-8")


def _merge_content_types(orig_data: bytes, new_data: bytes) -> bytes:
    """Merge [Content_Types].xml.

    Starts from the original (which registers types for all original parts:
    images, drawings, charts, …) and adds any new Override/Default entries
    from openpyxl (e.g. a newly added sheet). This ensures that files
    restored from the original ZIP still have registered content types.
    """
    default_tag = f"{{{_CT_NS}}}Default"
    override_tag = f"{{{_CT_NS}}}Override"

    orig_root = ET.fromstring(orig_data)
    new_root = ET.fromstring(new_data)

    orig_extensions = {el.get("Extension") for el in orig_root.iter(default_tag)}
    orig_partnames = {el.get("PartName") for el in orig_root.iter(override_tag)}

    for el in new_root.iter(default_tag):
        if el.get("Extension") not in orig_extensions:
            orig_root.append(el)
    for el in new_root.iter(override_tag):
        if el.get("PartName") not in orig_partnames:
            orig_root.append(el)

    xml_str = ET.tostring(orig_root, encoding="unicode")
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_str
    ).encode("utf-8")


_WORKSHEET_RE = re.compile(r"^xl/worksheets/sheet\d+\.xml$")

_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_R_NS_DECL = f'xmlns:r="{_R_NS}"'


def _ensure_r_namespace(element_snippet: str) -> str:
    """Add xmlns:r declaration if the element uses r: prefix but lacks the declaration.

    openpyxl does not declare r: on the worksheet root element; it inlines
    xmlns:r on each child element that needs it (legacyDrawing, hyperlink).
    Drawing elements restored from the original must follow the same pattern.
    """
    if "r:" not in element_snippet or _R_NS_DECL in element_snippet:
        return element_snippet
    # Insert xmlns:r before the first attribute or before the closing />
    return re.sub(r"(<\w+)", rf"\1 {_R_NS_DECL}", element_snippet, count=1)


# Elements openpyxl silently drops from worksheet XML.
# They must be restored verbatim so Excel can locate drawings/images.
_DROPPED_ELEMENT_TAGS = ("drawing", "legacyDrawing", "picture")


def _extract_drawing_elements(sheet_xml: bytes) -> list[str]:
    """Return raw XML snippets of drawing/legacyDrawing/picture elements."""
    tag_pattern = "|".join(_DROPPED_ELEMENT_TAGS)
    return re.findall(
        rf"<(?:{tag_pattern})(?:\s[^>]*)?\s*/>",
        sheet_xml.decode("utf-8", errors="replace"),
    )


def _inject_drawing_elements(new_xml: bytes, elements: list[str]) -> bytes:
    """Inject missing drawing/picture elements into openpyxl's worksheet XML.

    Inserts them right before </worksheet> so they appear in the correct
    position per OOXML spec (after sheetData, pageMargins, etc.).
    Drawing elements reference relationships by rId; the companion .rels
    merge ensures those rIds still resolve to the correct targets.

    legacyDrawing and picture are limited to one per sheet; openpyxl already
    writes its own version of these, so we never duplicate them.
    """
    if not elements:
        return new_xml

    text = new_xml.decode("utf-8", errors="replace")
    tag_pattern = "|".join(_DROPPED_ELEMENT_TAGS)

    # Tags where at most one element per sheet is allowed (openpyxl already wrote one)
    _SINGLE_INSTANCE_TAGS = {"legacyDrawing", "picture"}
    existing_tags = set(re.findall(rf"<({tag_pattern})[\s/>]", text))
    # Support both single- and double-quoted r:id attributes to avoid
    # duplicate injection when the original XML used single quotes.
    existing_rids = set(
        re.findall(rf"""(?:{tag_pattern})\s+[^>]*r:id=["']([^"']+)["']""", text)
    )

    to_inject = []
    for snippet in elements:
        tag_match = re.match(r"<(\w+)", snippet)
        if not tag_match:
            continue
        tag = tag_match.group(1)
        # Skip if this single-instance tag is already present in new XML
        if tag in _SINGLE_INSTANCE_TAGS and tag in existing_tags:
            continue
        rid_match = re.search(r"""r:id=["']([^"']+)["']""", snippet)
        rid = rid_match.group(1) if rid_match else None
        if rid is None or rid not in existing_rids:
            to_inject.append(snippet)

    if not to_inject:
        return new_xml

    injected_block = "\n".join(_ensure_r_namespace(s) for s in to_inject)
    text = text.replace("</worksheet>", f"{injected_block}</worksheet>", 1)
    return text.encode("utf-8")


def _patch_worksheet_xml(orig_data: bytes, new_data: bytes) -> bytes:
    """Restore drawing references dropped by openpyxl from worksheet XML."""
    orig_elements = _extract_drawing_elements(orig_data)
    return _inject_drawing_elements(new_data, orig_elements)


def _merge_workbook_zips(
    original_path: Path, new_buf: io.BytesIO, out_path: Path
) -> None:
    """Build a merged ZIP that preserves ALL original content.

    Algorithm:
    - Take every file from openpyxl's output (updated cells, styles, workbook XML).
    - For worksheet XML files: restore drawing/legacyDrawing/picture elements
      that openpyxl silently drops (images, charts, VML shapes).
    - For .rels files present in both ZIPs: merge relationship entries so
      openpyxl never silently drops drawing/image/chart references.
    - For [Content_Types].xml: merge from original + add new openpyxl entries.
    - For any file present in the original but absent from openpyxl's output
      (images in xl/media/, drawings in xl/drawings/, charts, pivot tables,
      ActiveX, …): restore it verbatim from the original.
    """
    new_buf.seek(0)
    with (
        zipfile.ZipFile(original_path, "r") as orig_zip,
        zipfile.ZipFile(new_buf, "r") as new_zip,
        zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as out_zip,
    ):
        orig_names = set(orig_zip.namelist())
        new_names = set(new_zip.namelist())

        for name in new_zip.namelist():
            data = new_zip.read(name)
            if name in orig_names:
                try:
                    if name == "[Content_Types].xml":
                        data = _merge_content_types(orig_zip.read(name), data)
                    elif name.endswith(".rels"):
                        # workbook.xml.rels must use openpyxl's rId numbering as
                        # the base because openpyxl rewrites workbook.xml with new
                        # rIds. Using the original as base would map those new rIds
                        # to wrong targets (theme, styles, vbaProject instead of
                        # the actual worksheet files).
                        prefer_new = name == "xl/_rels/workbook.xml.rels"
                        data = _merge_rels(
                            orig_zip.read(name), data, prefer_new=prefer_new
                        )
                    elif _WORKSHEET_RE.match(name):
                        # openpyxl drops <drawing>, <legacyDrawing>, <picture>
                        # elements from sheet XML; restore them from the original.
                        data = _patch_worksheet_xml(orig_zip.read(name), data)
                except Exception as exc:
                    warnings.warn(
                        f"Failed to merge ZIP entry {name!r}: {exc}. "
                        "Falling back to openpyxl's version; some content may be lost.",
                        stacklevel=3,
                    )
            out_zip.writestr(new_zip.getinfo(name), data)

        # Restore entries openpyxl dropped entirely
        for name in orig_names - new_names:
            out_zip.writestr(orig_zip.getinfo(name), orig_zip.read(name))


def safe_save_workbook(workbook: Workbook, path: str | Path) -> None:
    """Save workbook with zero data loss.

    Strategy:
    1. Back up original to <file>.bak.
    2. Save openpyxl's output to an in-memory buffer.
    3. Merge original ZIP and openpyxl buffer:
       - .rels files are merged (all original relationships preserved).
       - [Content_Types].xml is merged (original types preserved).
       - Files openpyxl dropped (images, drawings, charts, VBA, …) are
         restored verbatim from the original.
    4. Write merged output to a temp file in the same directory.
    5. Atomically rename temp → original (os.replace; POSIX atomic).

    If step 5 fails the original is untouched. The .bak copy provides an
    additional recovery path in case openpyxl's serialisation itself is lossy.
    """
    target = Path(path) if isinstance(path, str) else path
    backup = target.with_suffix(target.suffix + ".bak")
    shutil.copy2(target, backup)

    buf = io.BytesIO()
    workbook.save(buf)

    fd, tmp_path_str = tempfile.mkstemp(dir=target.parent, suffix=".tmp")
    try:
        os.close(fd)
        tmp_path = Path(tmp_path_str)
        _merge_workbook_zips(target, buf, tmp_path)
        os.replace(tmp_path, target)
        # Clean up backup only after successful atomic replace.
        try:
            backup.unlink()
        except OSError:
            pass  # non-critical; stale .bak is harmless
    except Exception:
        try:
            os.unlink(tmp_path_str)
        except OSError:
            pass
        raise
