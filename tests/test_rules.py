from __future__ import annotations

import pytest
from openpyxl import Workbook

from spreadsheet_tools.cleaner import build_merge_lookup
from spreadsheet_tools.rules import apply_rule


def _sheet_with_values(cells: dict[str, object]):
    wb = Workbook()
    ws = wb.active
    for addr, value in cells.items():
        ws[addr] = value
    return ws


def test_not_empty_passes_with_value() -> None:
    ws = _sheet_with_values({"A1": "ok"})
    result = apply_rule("not-empty:A1", ws)
    assert result.passed
    assert result.addresses == ["A1"]


def test_not_empty_fails_when_blank() -> None:
    ws = _sheet_with_values({})
    result = apply_rule("not-empty:A1", ws)
    assert not result.passed


def test_not_empty_range_flags_empty_cells() -> None:
    ws = _sheet_with_values({"A1": "x", "B1": None})
    result = apply_rule("not-empty-range:A1:B1", ws)
    assert not result.passed
    assert "B1" in result.addresses


def test_price_min_max() -> None:
    ws = _sheet_with_values({"A1": 10, "B1": 20})
    assert apply_rule("price-min-max:A1:B1", ws).passed
    ws["B1"] = 5
    assert not apply_rule("price-min-max:A1:B1", ws).passed


def test_name_matches_desc() -> None:
    ws = _sheet_with_values(
        {
            "A1": "Ventas regionales",
            "B1": "Estrategia de ventas regionales para el mercado local",
        }
    )
    assert apply_rule("name-matches-desc:A1:B1", ws).passed


def test_numeric_range() -> None:
    ws = _sheet_with_values({"C3": 15})
    assert apply_rule("numeric-range:C3:10:20", ws).passed
    ws["C3"] = 5
    assert not apply_rule("numeric-range:C3:10:20", ws).passed


def test_string_contains_case_insensitive() -> None:
    ws = _sheet_with_values({"D4": "Hello World"})
    assert apply_rule("string-contains:D4:world", ws).passed
    assert not apply_rule("string-contains:D4:missing", ws).passed


def test_no_generic_name() -> None:
    ws = _sheet_with_values({"E5": "Estrategia de marketing 3"})
    assert not apply_rule("no-generic-name:E5", ws).passed
    ws["E5"] = "Plan comercial Q1"
    assert apply_rule("no-generic-name:E5", ws).passed


def test_unknown_rule_raises() -> None:
    ws = _sheet_with_values({})
    with pytest.raises(ValueError, match="Unknown rule"):
        apply_rule("bogus:A1", ws)


def test_merged_slave_reads_master_value() -> None:
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:B1")
    ws["A1"] = "merged"
    lookup = build_merge_lookup(ws)
    result = apply_rule("not-empty:B1", ws, merge_lookup=lookup)
    assert result.passed
