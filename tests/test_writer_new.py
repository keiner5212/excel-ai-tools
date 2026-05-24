"""Integration tests for new writer commands.

Each test creates a temporary workbook, applies the new operation, and verifies
the result by re-reading the file. No raw openpyxl calls outside of setup
fixtures — all writes go through the CLI writer functions.
"""
from __future__ import annotations

import gc
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from spreadsheet_tools.writer import (
    add_sheet,
    batch_edit,
    batch_set_dimensions,
    create_empty_workbook,
    format_range,
    freeze_panes,
    merge_cells,
    rename_sheet,
    set_column_width,
    set_row_height,
    set_tab_color,
    toggle_auto_filter,
    unfreeze_panes,
    unmerge_cells,
)
from spreadsheet_tools.utils import validate_cell_range, resolve_output_path, open_workbook


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def tmp_workbook(tmp_path: Path) -> Path:
    """Return path to a minimal temporary workbook with one sheet."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Workbook must have an active sheet"
    ws.title = "Sheet1"
    ws["A1"] = "Title"
    ws["B2"] = 42
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# validate_cell_range
# ---------------------------------------------------------------------------


def test_validate_cell_range_valid() -> None:
    assert validate_cell_range("A1:E1") == ("A1", "E1")
    assert validate_cell_range(" b2:d10 ") == ("B2", "D10")


def test_validate_cell_range_invalid() -> None:
    with pytest.raises(ValueError, match="Invalid cell range"):
        validate_cell_range("A1")
    with pytest.raises(ValueError, match="Invalid cell range"):
        validate_cell_range("A0:B1")


# ---------------------------------------------------------------------------
# resolve_output_path
# ---------------------------------------------------------------------------


def test_resolve_output_path_bare_goes_to_workspace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    result = resolve_output_path("file.xlsx")
    assert result == Path("workspace") / "file.xlsx"
    assert (tmp_path / "workspace").is_dir()


def test_resolve_output_path_explicit_dir_unchanged(tmp_path: Path) -> None:
    result = resolve_output_path(str(tmp_path / "sub" / "file.xlsx"))
    assert result == tmp_path / "sub" / "file.xlsx"


# ---------------------------------------------------------------------------
# merge_cells / unmerge_cells
# ---------------------------------------------------------------------------


def test_merge_cells_saves_and_readable(tmp_workbook: Path) -> None:
    result = merge_cells(str(tmp_workbook), sheet_name="Sheet1", cell_range="A1:C1")
    assert result["saved"] is True
    assert result["merged"] == "A1:C1"
    assert result["master"] == "A1"
    wb = load_workbook(tmp_workbook)
    ranges = [str(r) for r in wb["Sheet1"].merged_cells.ranges]
    assert "A1:C1" in ranges


def test_unmerge_cells(tmp_workbook: Path) -> None:
    merge_cells(str(tmp_workbook), sheet_name="Sheet1", cell_range="A1:C1")
    result = unmerge_cells(str(tmp_workbook), sheet_name="Sheet1", cell_range="A1:C1")
    assert result["saved"] is True
    wb = load_workbook(tmp_workbook)
    ranges = [str(r) for r in wb["Sheet1"].merged_cells.ranges]
    assert "A1:C1" not in ranges


def test_merge_cells_invalid_range(tmp_workbook: Path) -> None:
    with pytest.raises(ValueError, match="Invalid cell range"):
        merge_cells(str(tmp_workbook), sheet_name="Sheet1", cell_range="A1")


# ---------------------------------------------------------------------------
# set_column_width / set_row_height / batch_set_dimensions
# ---------------------------------------------------------------------------


def test_set_column_width(tmp_workbook: Path) -> None:
    result = set_column_width(str(tmp_workbook), sheet_name="Sheet1", col="B", width=20.5)
    assert result["saved"] is True
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].column_dimensions["B"].width == 20.5


def test_set_column_width_rejects_zero(tmp_workbook: Path) -> None:
    with pytest.raises(ValueError, match="must be > 0"):
        set_column_width(str(tmp_workbook), sheet_name="Sheet1", col="A", width=0)


def test_set_row_height(tmp_workbook: Path) -> None:
    result = set_row_height(str(tmp_workbook), sheet_name="Sheet1", row=1, height=40)
    assert result["saved"] is True
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].row_dimensions[1].height == 40


def test_set_row_height_rejects_row_zero(tmp_workbook: Path) -> None:
    with pytest.raises(ValueError, match="Row number must be >= 1"):
        set_row_height(str(tmp_workbook), sheet_name="Sheet1", row=0, height=20)


def test_batch_set_dimensions(tmp_workbook: Path) -> None:
    result = batch_set_dimensions(
        str(tmp_workbook),
        sheet_name="Sheet1",
        column_widths=[{"col": "A", "width": 8}, {"col": "B", "width": 15}],
        row_heights=[{"row": 1, "height": 30}],
    )
    assert result["saved"] is True
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].column_dimensions["A"].width == 8
    assert wb["Sheet1"].column_dimensions["B"].width == 15
    assert wb["Sheet1"].row_dimensions[1].height == 30


def test_batch_set_dimensions_requires_at_least_one(tmp_workbook: Path) -> None:
    with pytest.raises(ValueError, match="At least one"):
        batch_set_dimensions(str(tmp_workbook), sheet_name="Sheet1")


# ---------------------------------------------------------------------------
# freeze_panes / unfreeze_panes
# ---------------------------------------------------------------------------


def test_freeze_panes(tmp_workbook: Path) -> None:
    result = freeze_panes(str(tmp_workbook), sheet_name="Sheet1", cell="B2")
    assert result["freeze_panes"] == "B2"
    assert result["saved"] is True
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].freeze_panes == "B2"


def test_unfreeze_panes(tmp_workbook: Path) -> None:
    freeze_panes(str(tmp_workbook), sheet_name="Sheet1", cell="B2")
    result = unfreeze_panes(str(tmp_workbook), sheet_name="Sheet1")
    assert result["freeze_panes"] is None
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].freeze_panes is None


# ---------------------------------------------------------------------------
# add_sheet / rename_sheet
# ---------------------------------------------------------------------------


def test_add_sheet(tmp_workbook: Path) -> None:
    result = add_sheet(str(tmp_workbook), sheet_name="NewSheet")
    assert result["added"] is True
    wb = load_workbook(tmp_workbook)
    assert "NewSheet" in wb.sheetnames


def test_add_sheet_duplicate_raises(tmp_workbook: Path) -> None:
    with pytest.raises(ValueError, match="already exists"):
        add_sheet(str(tmp_workbook), sheet_name="Sheet1")


def test_rename_sheet(tmp_workbook: Path) -> None:
    result = rename_sheet(str(tmp_workbook), old_name="Sheet1", new_name="Renamed")
    assert result["new_name"] == "Renamed"
    wb = load_workbook(tmp_workbook)
    assert "Renamed" in wb.sheetnames
    assert "Sheet1" not in wb.sheetnames


def test_rename_sheet_not_found_raises(tmp_workbook: Path) -> None:
    with pytest.raises(ValueError, match="not found"):
        rename_sheet(str(tmp_workbook), old_name="Ghost", new_name="X")


# ---------------------------------------------------------------------------
# format_range
# ---------------------------------------------------------------------------


def test_format_range_applies_style(tmp_workbook: Path) -> None:
    result = format_range(
        str(tmp_workbook),
        sheet_name="Sheet1",
        cell_range="A1:B2",
        style={"font": {"bold": True}},
    )
    assert result["cells_styled"] == 4
    assert result["saved"] is True
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].font.bold is True
    assert wb["Sheet1"]["B2"].font.bold is True


def test_format_range_skips_slaves(tmp_workbook: Path) -> None:
    merge_cells(str(tmp_workbook), sheet_name="Sheet1", cell_range="A1:C1")
    result = format_range(
        str(tmp_workbook),
        sheet_name="Sheet1",
        cell_range="A1:C1",
        style={"font": {"italic": True}},
    )
    # Only master cell (A1) styled; B1 and C1 are slaves → 1 cell styled
    assert result["cells_styled"] == 1


# ---------------------------------------------------------------------------
# set_tab_color
# ---------------------------------------------------------------------------


def test_set_tab_color(tmp_workbook: Path) -> None:
    result = set_tab_color(str(tmp_workbook), sheet_name="Sheet1", color="1F4E79")
    assert result["tab_color"] == "FF1F4E79"
    wb = load_workbook(tmp_workbook)
    tab_color = wb["Sheet1"].sheet_properties.tabColor
    assert tab_color is not None
    assert tab_color.rgb == "FF1F4E79"


def test_set_tab_color_invalid(tmp_workbook: Path) -> None:
    with pytest.raises(ValueError, match="Color must be"):
        set_tab_color(str(tmp_workbook), sheet_name="Sheet1", color="ZZZ")


# ---------------------------------------------------------------------------
# toggle_auto_filter
# ---------------------------------------------------------------------------


def test_auto_filter_enable(tmp_workbook: Path) -> None:
    result = toggle_auto_filter(
        str(tmp_workbook), sheet_name="Sheet1", cell_range="A1:B1"
    )
    assert result["auto_filter"] == "A1:B1"
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].auto_filter.ref == "A1:B1"


def test_auto_filter_clear(tmp_workbook: Path) -> None:
    toggle_auto_filter(str(tmp_workbook), sheet_name="Sheet1", cell_range="A1:B1")
    result = toggle_auto_filter(
        str(tmp_workbook), sheet_name="Sheet1", cell_range=None
    )
    assert result["auto_filter"] is None
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"].auto_filter.ref is None


# ---------------------------------------------------------------------------
# Regression: Fix 1 — create_empty_workbook returns absolute path
# ---------------------------------------------------------------------------


def test_create_empty_workbook_returns_absolute_path(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Returned 'file' path must be absolute so callers can reuse it directly."""
    monkeypatch.chdir(tmp_path)
    result = create_empty_workbook("report.xlsx")
    returned_path = Path(result["file"])
    assert returned_path.is_absolute(), "create_empty_workbook must return absolute path"
    assert returned_path.exists()
    assert result["sheet"] == "Sheet1"
    assert result["created"] is True


def test_create_empty_workbook_custom_sheet(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.chdir(tmp_path)
    result = create_empty_workbook("out.xlsx", sheet_name="Data")
    wb = load_workbook(result["file"])
    assert "Data" in wb.sheetnames


# ---------------------------------------------------------------------------
# Regression: Fix 2 — get_sheet coerces literal "null" string to None
# ---------------------------------------------------------------------------


def test_get_sheet_null_string_uses_active_sheet(tmp_workbook: Path) -> None:
    """Passing sheet_name='null' (string) must select the active sheet, not fail."""
    from spreadsheet_tools.utils import get_sheet, open_workbook

    workbook = open_workbook(str(tmp_workbook), read_only=False, data_only=False)
    try:
        sheet = get_sheet(workbook, "null")
        assert sheet is not None
        assert sheet.title == "Sheet1"
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Regression: Fix 3 — batch_edit validates edit structure early
# ---------------------------------------------------------------------------


def test_batch_edit_rejects_non_dict_entry(tmp_workbook: Path) -> None:
    """batch_edit must raise ValueError when an entry is not a dict."""
    with pytest.raises(ValueError, match="must be a dict"):
        batch_edit(str(tmp_workbook), edits=["A1"], sheet_name="Sheet1")  # type: ignore[list-item]


def test_batch_edit_rejects_missing_cell_key(tmp_workbook: Path) -> None:
    """batch_edit must raise ValueError when 'cell' key is absent."""
    with pytest.raises(ValueError, match="missing required 'cell'"):
        batch_edit(str(tmp_workbook), edits=[{"value": "x"}], sheet_name="Sheet1")


def test_batch_edit_rejects_invalid_cell_address(tmp_workbook: Path) -> None:
    """batch_edit must raise ValueError for malformed cell addresses."""
    with pytest.raises(ValueError):
        batch_edit(str(tmp_workbook), edits=[{"cell": "ZZ0"}], sheet_name="Sheet1")


def test_batch_edit_applies_valid_edits(tmp_workbook: Path) -> None:
    """batch_edit must write multiple cells atomically."""
    result = batch_edit(
        str(tmp_workbook),
        edits=[{"cell": "C3", "value": "hello"}, {"cell": "D4", "value": 99}],
        sheet_name="Sheet1",
    )
    assert result["saved"] is True
    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["C3"].value == "hello"
    assert wb["Sheet1"]["D4"].value == 99


# ---------------------------------------------------------------------------
# Regression: Fix 4 — vba_archive ZipFile leak (PytestUnraisableExceptionWarning)
# ---------------------------------------------------------------------------


def test_vba_archive_closed_after_write(tmp_workbook: Path) -> None:
    """Writer functions must leave no open ZipFile after returning.

    If vba_archive is left open, GC may trigger ZipFile.__del__ with a
    closed BytesIO, raising 'I/O operation on closed file'.
    Running gc.collect() here with -W error would catch that regression.
    """
    merge_cells(str(tmp_workbook), sheet_name="Sheet1", cell_range="A1:B1")
    gc.collect()  # triggers __del__ on any leaked ZipFile objects


def test_vba_archive_closed_on_error_path(tmp_workbook: Path) -> None:
    """Error paths (ValueError before save) must also close vba_archive."""
    with pytest.raises(ValueError):
        add_sheet(str(tmp_workbook), sheet_name="Sheet1")  # duplicate → raises
    gc.collect()  # must not trigger ZipFile.__del__ error


def test_open_workbook_vba_archive_closed_on_close(tmp_workbook: Path) -> None:
    """open_workbook().close() must close vba_archive to prevent GC errors."""
    workbook = open_workbook(str(tmp_workbook), read_only=False, data_only=False)
    workbook.close()
    assert workbook.vba_archive is None or workbook.vba_archive.fp is None
    gc.collect()
